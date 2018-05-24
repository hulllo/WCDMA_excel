from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

class Excel():
    
    def __init__(self, filename):
        self.filename = filename
        self.redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
        self.whiteFill = PatternFill(start_color='FFFFFFFF',
                   end_color='FFFFFFFF',
                   fill_type='solid')    
    def writeto(self, list_): 
        wb=load_workbook(self.filename)
        sheetnames  = wb.sheetnames
        for x in list_:
            sheet_name = x[0]
            row = x[1]
            col = x[2]
            value = x[3]
            if sheet_name not in sheetnames:
                print('sheet_name do not exist')
                return False
            else:
                ws = wb[sheet_name]
                ws.cell(row,col).value = value
                if value == 'Failed':
                    ws.cell(row, col).fill = self.redFill
                else:
                    ws.cell(row, col).fill = self.whiteFill
        wb.save('./a_changed.xlsx')

def main():
    excel = Excel('./a.xlsx')
    excel.writeto('LTEB3', 12, 6, '23')
            
            
if __name__ == '__main__':
    main()
